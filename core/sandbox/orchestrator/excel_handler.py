"""
Excel Handler - Scripts for Excel operations in sandbox containers

This module contains Python scripts that are executed inside sandbox containers
to perform Excel operations using openpyxl and formulas libraries.
"""

def get_read_excel_script(file_path: str, sheet_index: int) -> str:
    """
    Generate Python script to read Excel file with formulas and values.

    Args:
        file_path: Absolute path to Excel file in container
        sheet_index: Index of sheet to read (0-based)

    Returns:
        Python script as string
    """
    return f"""
import openpyxl
import json
import base64

try:
    # Read file as binary
    with open({repr(file_path)}, 'rb') as f:
        excel_bytes = f.read()

    from io import BytesIO

    # Load workbook twice: once for values, once for formulas
    wb_values = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)
    wb_formulas = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=False)

    sheet_names = wb_formulas.sheetnames

    if {sheet_index} >= len(sheet_names):
        print(json.dumps({{"error": "Invalid sheet index"}}))
    else:
        ws_values = wb_values.worksheets[{sheet_index}]
        ws_formulas = wb_formulas.worksheets[{sheet_index}]

        max_row = ws_formulas.max_row
        max_col = ws_formulas.max_column

        data = []
        formulas = []

        for row_idx in range(max_row):
            row_data = []
            row_formulas = []

            for col_idx in range(max_col):
                # Get calculated value from data_only workbook
                cell_value = ws_values.cell(row=row_idx + 1, column=col_idx + 1)
                value = cell_value.value if cell_value.value is not None else ""

                # Get formula from formula workbook
                cell_formula = ws_formulas.cell(row=row_idx + 1, column=col_idx + 1)
                formula = None

                if cell_formula.data_type == 'f':
                    formula_str = str(cell_formula.value)
                    # Remove leading '=' if present
                    formula = formula_str[1:] if formula_str.startswith('=') else formula_str

                row_data.append(value)
                row_formulas.append(formula)

            data.append(row_data)
            formulas.append(row_formulas)

        # Get column widths
        column_widths = {{}}
        for col_idx in range(max_col):
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_idx + 1)
            width = ws_formulas.column_dimensions[col_letter].width
            if width:
                column_widths[col_idx] = int(width * 7)

        result = {{
            "sheet_names": sheet_names,
            "data": data,
            "formulas": formulas,
            "column_widths": column_widths
        }}
        print(json.dumps(result))

except Exception as e:
    import traceback
    print(json.dumps({{"error": str(e), "traceback": traceback.format_exc()}}))
"""


def get_update_cell_script(
    file_path: str,
    sheet_index: int,
    row: int,
    col: int,
    value: str = None,
    formula: str = None
) -> str:
    """
    Generate Python script to update a cell with value or formula.

    This script:
    1. Updates the specified cell
    2. Recalculates ALL formulas in the entire workbook (cascade update)
    3. Returns all updated cell values and formulas

    Args:
        file_path: Absolute path to Excel file in container
        sheet_index: Index of sheet (0-based)
        row: Row index (0-based)
        col: Column index (0-based)
        value: Cell value (mutually exclusive with formula)
        formula: Cell formula without '=' (mutually exclusive with value)

    Returns:
        Python script as string that outputs JSON with:
        - evaluated_value: value of the updated cell
        - updated_cells: dict of all cells that changed (for cascade update)
    """
    return f"""
import openpyxl
import json

try:
    # Load workbook
    wb = openpyxl.load_workbook({repr(file_path)}, data_only=False)

    if {sheet_index} >= len(wb.sheetnames):
        print(json.dumps({{"error": "Invalid sheet index"}}))
    else:
        ws = wb.worksheets[{sheet_index}]

        # Get cell (1-indexed in openpyxl)
        cell = ws.cell(row={row + 1}, column={col + 1})

        evaluated_value = None
        is_formula = False
        sheet_name = ws.title  # Get sheet name before closing

        if {repr(formula)}:
            formula_str = {repr(formula)}
            if not formula_str.startswith('='):
                formula_str = '=' + formula_str
            cell.value = formula_str
            is_formula = True
        elif {repr(value)} is not None:
            val = {repr(value)}
            try:
                # Try to convert to number
                if '.' in str(val):
                    cell.value = float(val)
                else:
                    cell.value = int(val)
            except (ValueError, TypeError):
                cell.value = val
            evaluated_value = cell.value
        else:
            cell.value = None
            evaluated_value = ""

        # Save workbook
        wb.save({repr(file_path)})
        wb.close()

        # If formula, evaluate it using formulas library
        # Otherwise just return the value we already have
        if is_formula:
            try:
                import formulas
                import os

                # Read the entire Excel file and compile it
                xl_model = formulas.ExcelModel().loads({repr(file_path)}).finish()

                # Get cell reference in format: [filename.xlsx]SHEET1!A1
                # Note: formulas library converts sheet names to uppercase
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter({col + 1})
                filename = os.path.basename({repr(file_path)})
                cell_ref = "'" + "[" + filename + "]" + sheet_name.upper() + "'!" + col_letter + str({row + 1})

                # Calculate all formulas
                result = xl_model.calculate()

                # Get the evaluated value for our cell
                if cell_ref in result:
                    val = result[cell_ref]
                    # Extract value from Ranges object if needed
                    if hasattr(val, 'value'):
                        evaluated_value = val.value[0][0] if val.value and val.value[0] else formula_str
                    else:
                        evaluated_value = val
                else:
                    # Cell might have been calculated under different name, just show formula
                    evaluated_value = formula_str
            except Exception as e:
                # If formulas library fails (not installed or error), just show the formula
                import traceback
                print("Note: Formula not evaluated: " + str(e), file=__import__('sys').stderr)
                # Don't fail, just show the formula text as fallback
                evaluated_value = formula_str

        # Recalculate ALL cells in the entire workbook for cascade updates
        updated_cells = {{}}
        try:
            import formulas
            import os

            # Read and calculate all formulas
            xl_model = formulas.ExcelModel().loads({repr(file_path)}).finish()
            filename = os.path.basename({repr(file_path)})
            result_dict = xl_model.calculate()

            # Extract all calculated values for this sheet
            for key, val in result_dict.items():
                # Filter for current sheet and valid cell references
                if filename in key and sheet_name.upper() in key and "'!" in key:
                    try:
                        # Parse the cell reference to get row/col
                        cell_part = key.split("'!")[-1]
                        # Extract column letter and row number
                        import re
                        match = re.match(r'([A-Z]+)(\\d+)', cell_part)
                        if match:
                            from openpyxl.utils import column_index_from_string
                            cell_col_letter = match.group(1)
                            cell_row_num = int(match.group(2))
                            cell_col_idx = column_index_from_string(cell_col_letter) - 1
                            cell_row_idx = cell_row_num - 1

                            # Extract value from Ranges object
                            if hasattr(val, 'value') and val.value and len(val.value) > 0:
                                cell_value = val.value[0][0] if len(val.value[0]) > 0 else None
                                if cell_value is not None:
                                    updated_cells[f"{{cell_row_idx}}_{{cell_col_idx}}"] = cell_value
                    except:
                        pass
        except Exception as calc_error:
            # If formulas library fails, just continue without cascade updates
            import traceback
            print("Note: Cascade recalculation failed: " + str(calc_error), file=__import__('sys').stderr)

        result = {{
            "evaluated_value": evaluated_value,
            "updated_cells": updated_cells
        }}
        print(json.dumps(result))

except Exception as e:
    import traceback
    print(json.dumps({{"error": str(e), "traceback": traceback.format_exc()}}))
"""


def get_batch_update_cells_script(
    file_path: str,
    sheet_index: int,
    updates: list  # List of dicts: [{"row": 0, "col": 0, "value": "x", "formula": None}, ...]
) -> str:
    """
    Generate Python script to update multiple cells in batch.

    This is MUCH faster than calling update_cell multiple times because:
    1. It opens the workbook only once
    2. It updates all cells in memory
    3. It saves the workbook only once
    4. It recalculates formulas only once

    Args:
        file_path: Absolute path to Excel file in container
        sheet_index: Index of sheet (0-based)
        updates: List of cell updates, each with row, col, value/formula

    Returns:
        Python script as string that outputs JSON with all updated cell values
    """
    # Serialize updates list to JSON string for embedding in script
    import json
    updates_json = json.dumps(updates)

    return f"""
import openpyxl
import json

try:
    # Load workbook once
    wb = openpyxl.load_workbook({repr(file_path)}, data_only=False)

    if {sheet_index} >= len(wb.sheetnames):
        print(json.dumps({{"error": "Invalid sheet index"}}))
    else:
        ws = wb.worksheets[{sheet_index}]
        sheet_name = ws.title

        # Parse updates from JSON
        updates = {updates_json}

        # Apply all updates to workbook in memory
        for update in updates:
            row = update["row"]
            col = update["col"]
            value = update.get("value")
            formula = update.get("formula")

            cell = ws.cell(row=row + 1, column=col + 1)

            if formula:
                formula_str = formula if formula.startswith('=') else '=' + formula
                cell.value = formula_str
            elif value is not None:
                try:
                    # Try to convert to number
                    if '.' in str(value):
                        cell.value = float(value)
                    else:
                        cell.value = int(value)
                except (ValueError, TypeError):
                    cell.value = value
            else:
                cell.value = None

        # Save workbook once (instead of N times)
        wb.save({repr(file_path)})
        wb.close()

        # Recalculate ALL formulas once using formulas library
        updated_cells = {{}}
        try:
            import formulas
            import os

            xl_model = formulas.ExcelModel().loads({repr(file_path)}).finish()
            filename = os.path.basename({repr(file_path)})
            result_dict = xl_model.calculate()

            # Extract all calculated values for this sheet
            for key, val in result_dict.items():
                if filename in key and sheet_name.upper() in key and "'!" in key:
                    try:
                        cell_part = key.split("'!")[-1]
                        import re
                        match = re.match(r'([A-Z]+)(\\d+)', cell_part)
                        if match:
                            from openpyxl.utils import column_index_from_string
                            cell_col_letter = match.group(1)
                            cell_row_num = int(match.group(2))
                            cell_col_idx = column_index_from_string(cell_col_letter) - 1
                            cell_row_idx = cell_row_num - 1

                            if hasattr(val, 'value') and val.value and len(val.value) > 0:
                                cell_value = val.value[0][0] if len(val.value[0]) > 0 else None
                                if cell_value is not None:
                                    updated_cells[f"{{cell_row_idx}}_{{cell_col_idx}}"] = cell_value
                    except:
                        pass
        except Exception as calc_error:
            import traceback
            print("Note: Cascade recalculation failed: " + str(calc_error), file=__import__('sys').stderr)

        result = {{
            "success": True,
            "updated_cells": updated_cells,
            "count": len(updates)
        }}
        print(json.dumps(result))

except Exception as e:
    import traceback
    print(json.dumps({{"error": str(e), "traceback": traceback.format_exc()}}))
"""
